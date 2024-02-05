import pandas as pd
import glob
import os

from openpyxl.utils import get_column_letter

file_path = r'C:\Users\jstco\Downloads\PythonML-main\PythonML-main\chapter04'

m_store = pd.read_csv(file_path + '/m_store.csv')
m_area = pd.read_csv(file_path + '/m_area.csv')

tbl_order_file = os.path.join(file_path, 'tbl_order_*.csv')
tbl_order_files = glob.glob(tbl_order_file)
# print(tbl_order_files)

order_all = pd.DataFrame()
for file in tbl_order_files:
    order_tmp = pd.read_csv(file)
    # print(f'{file} : {len(order_tmp)}')
    order_all = pd.concat([order_all, order_tmp], ignore_index = True)

order_all = order_all.loc[order_all['store_id'] != 999]

order_all = pd.merge(order_all, m_store, on = 'store_id', how = 'left')
order_all = pd.merge(order_all, m_area, on = 'area_cd', how = 'left')

takeout_mapping = {0: 'delevery', 1: 'takeout'}
status_mapping = {0: '주문접수', 1: '결제완료', 2: '배달완료', 9: '주문취소'}

order_all['takeout_name'] = order_all['takeout_flag'].map(takeout_mapping)
order_all['status_name'] = order_all['status'].map(status_mapping)

order_all.loc[:, 'order_date'] = pd.to_datetime(order_all['order_accept_date']).dt.date

# print(order_all.head())

import openpyxl as op

# wb = op.Workbook()
# ws = wb['Sheet']
# ws.cell(1, 1).value = 'Writing test'
# wb.save('test.xlsx')
# wb.close()
#
# wb = op.load_workbook('test.xlsx', read_only = True)
# ws = wb['Sheet']
# print(ws.cell(1, 1).value)
# wb.close()

## 테스트 데이터 준비
store_id = 1
store_df = order_all.loc[order_all['store_id'] == store_id].copy()
store_name = store_df['store_name'].unique()[0]
store_sales_total = store_df.loc[store_df['status'].isin([1, 2])]['total_amount'].sum()
store_sales_takeout = store_df.loc[store_df['status'] == 1]['total_amount'].sum()
store_sales_delivery = store_df.loc[store_df['status'] == 2]['total_amount'].sum()

# print(f'매출액 확인 : {store_sales_total} = ' f'{store_sales_takeout + store_sales_delivery}')

output_df = store_df[['order_accept_date', 'customer_id', 'total_amount', 'takeout_name', 'status_name']]
# print(output_df.head())

## 특정 지점 데이터 익스포트하기
from openpyxl.utils.dataframe import dataframe_to_rows

store_title = f'{store_id}_{store_name}'

wb = op.Workbook()
ws = wb.active
ws.title = store_title

ws.cell(1, 1).value = f'{store_title} 매출 데이터'

rows = dataframe_to_rows(output_df, index = False, header = True)

row_start = 3
col_start = 2

for row_no, row in enumerate(rows, row_start):
    for col_no, value in enumerate(row, col_start):
        ws.cell(row_no, col_no).value = value

filename = f'{store_title}.xlsx'
wb.save(filename)
wb.close()

## 정리해서 출력하기
from openpyxl.styles import PatternFill, Border, Side, Font

# op.load_workbook(filename)
# # print(store_title)
# ws = wb[store_title]
#
# side = Side(style = 'thin', color = '008080')
# border = Border(top = side, bottom = side, left = side, right = side)
#
# for row in ws:
#     for cell in row:
#         if ws[cell.coordinate].value:
#             ws[cell.coordinate].border = border
#
# ws.cell(1, 1).font = Font(bold = True, color = '008080')

# 열 데이터 설정
# titles = ['주문접수일시', '고객아이디', '수입총액', '주문타입', '주문상태']
# pattern_fill = PatternFill(patternType = 'solid', fgColor = '008080')
# font_style = Font(bold = True, color = 'FFFFFF')
#
# # 열 데이터 입력
# for col, title in enumerate(titles, 2):
#     cell = ws.cell(3, col)
#     cell.fill = pattern_fill
#     cell.value = title
#     cell.font = font_style
#
# # 열 너비 설정
# column_widths = [20, 20, 12, 12, 12, 12]
# for col, width in enumerate(column_widths, 1):
#     ws.column_dimensions[get_column_letter(col)].width = width

# wb.save(filename)
# print(filename)
# wb.close()
